from flask import Flask, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from config import Config
import logging
from model import db, Faculty, Attendance, Venue, VenueAllocation
from datetime import datetime
from flask_cors import CORS
import random
import openpyxl
from io import BytesIO
import re
from flask import Response

app = Flask(__name__)

CORS(app, resources={
    r"/api/*": {
        "origins": ["http://localhost:5173"],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Authorization", "Content-Type"],
        "supports_credentials": True
    }
})

app.config.from_object(Config)
db.init_app(app)
jwt = JWTManager(app)
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Create tables only once at startup
with app.app_context():
    db.create_all()

@app.before_request
def handle_preflight():
    if request.method == "OPTIONS":
        response = app.make_response("")
        response.headers["Access-Control-Allow-Origin"] = "http://localhost:5173"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
        response.headers["Access-Control-Allow-Headers"] = "Authorization, Content-Type"
        response.headers["Access-Control-Max-Age"] = "3600"
        return response

@app.after_request
def after_request(response):
    response.headers["Access-Control-Allow-Origin"] = "http://localhost:5173"
    response.headers["Access-Control-Allow-Headers"] = "Authorization, Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    db.session.commit()
    db.session.remove()
    return response

@jwt.unauthorized_loader
def unauthorized_response(callback):
    logger.error(f"Unauthorized access: {callback}")
    return jsonify({'message': 'Missing or invalid token'}), 401

@jwt.invalid_token_loader
def invalid_token_response(callback):
    logger.error(f"Invalid token: {callback}")
    return jsonify({'message': 'Invalid token'}), 422

@jwt.expired_token_loader
def expired_token_response(jwt_header, jwt_payload):
    logger.error(f"Expired token: {jwt_payload}")
    return jsonify({'message': 'Token has expired'}), 401

# Login API
@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')

    faculty = Faculty.query.filter_by(email_id=email).first()
    if not faculty or faculty.mobile_number != password:
        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401

    access_token = create_access_token(identity=str(faculty.faculty_id))
    return jsonify({
        'success': True,
        'token': access_token,
        'is_admin': faculty.is_admin,
        'user': {'id': faculty.faculty_id, 'name': faculty.name, 'email': faculty.email_id}
    })

# Logout API (client-side token removal)
@app.route('/api/logout', methods=['POST'])
@jwt_required()
def logout():
    return jsonify({'success': True, 'message': 'Logged out successfully'})

@app.route('/api/current_user', methods=['GET'])
@jwt_required()
def get_current_user():
    faculty_id = get_jwt_identity()
    faculty = db.session.get(Faculty, faculty_id)
    if not faculty:
        return jsonify({'message': 'User not found'}), 404
    return jsonify({
        'id': faculty.faculty_id,
        'name': faculty.name,
        'email': faculty.email_id,
        'is_admin': faculty.is_admin
    })

# Mark Attendance
@app.route('/api/attendance', methods=['POST'])
@jwt_required()
def mark_attendance():
    data = request.get_json()
    date = datetime.strptime(data.get('date', datetime.now().strftime('%Y-%m-%d')), '%Y-%m-%d').date()
    allocation_id = data.get('allocation_id')  # Required
    rfid_tag = data.get('rfid_tag')

    if not allocation_id:
        return jsonify({'success': False, 'message': 'Allocation ID is required'}), 400

    if rfid_tag:
        if not re.match(r'^\d{10}$', rfid_tag):
            return jsonify({'success': False, 'message': 'RFID must be a 10-digit number'}), 400
        faculty = Faculty.query.filter_by(rfid_tag=rfid_tag).first()
        if not faculty:
            return jsonify({'success': False, 'message': 'Faculty not found with this RFID'}), 404
        faculty_id = faculty.faculty_id
    else:
        faculty_id = get_jwt_identity()
        faculty = db.session.get(Faculty, faculty_id)
        if not faculty:
            return jsonify({'success': False, 'message': 'Faculty not found'}), 404

    # Verify the allocation exists and matches the faculty and date
    allocation = db.session.get(VenueAllocation, allocation_id)
    if not allocation or allocation.faculty_id != faculty_id or allocation.date != date:
        return jsonify({'success': False, 'message': 'Invalid allocation for this faculty or date'}), 404

    # Check if attendance is already marked for this allocation
    attendance = Attendance.query.filter_by(allocation_id=allocation_id).first()
    if attendance:
        if attendance.is_present:
            return jsonify({'success': False, 'message': 'Attendance already marked for this allocation'}), 400
        attendance.is_present = True
    else:
        attendance = Attendance(
            faculty_id=faculty_id,
            allocation_id=allocation_id,
            date=date,
            is_present=True
        )
        db.session.add(attendance)

    db.session.commit()
    return jsonify({'success': True, 'message': 'Attendance marked successfully'})
# Get Attendance Records (Admin only)
@app.route('/api/attendance_records', methods=['GET'])
@jwt_required()
def get_attendance_records():
    faculty_id = get_jwt_identity()
    faculty = db.session.get(Faculty, faculty_id)
    if not faculty or not faculty.is_admin:
        return jsonify({'message': 'Unauthorized'}), 403

    date_filter = request.args.get('date', 'all')
    export_format = request.args.get('export', None)

    query = Attendance.query.join(Faculty).join(VenueAllocation, isouter=True)
    if date_filter != 'all':
        query = query.filter(Attendance.date == date_filter)

    records = query.all()
    data = [{
        'id': r.id,
        'faculty_id': r.faculty_id,
        'faculty_name': r.faculty.name,
        'rfid_tag': r.faculty.rfid_tag,
        'allocation_id': r.allocation_id,
        'venue_name': r.allocation.venue.name if r.allocation else 'N/A',
        'date': r.date.isoformat(),
        'is_present': r.is_present
    } for r in records]

    if export_format:
        if export_format == 'excel':
            import pandas as pd
            df = pd.DataFrame(data)
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            output.seek(0)
            return send_file(output, download_name=f'attendance_{date_filter}.xlsx', as_attachment=True)
        elif export_format == 'pdf':
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table as RLTable
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=letter)
            elements = []
            table_data = [['ID', 'Faculty', 'RFID', 'Venue', 'Date', 'Present']] + [
                [r['id'], r['faculty_name'], r['rfid_tag'], r['venue_name'], r['date'], 'Yes' if r['is_present'] else 'No']
                for r in data
            ]
            table = RLTable(table_data)
            elements.append(table)
            doc.build(elements)
            output.seek(0)
            return send_file(output, download_name=f'attendance_{date_filter}.pdf', as_attachment=True)

    return jsonify(data)

@app.route('/api/venues', methods=['GET'])
@jwt_required()
def get_venues():
    venues = Venue.query.all()
    return jsonify([{
        'venue_id': v.venue_id,
        'name': v.name,
        'location': v.location,
        'capacity': v.capacity
    } for v in venues])

@app.route('/api/venues', methods=['POST'])
@jwt_required()
def add_venue():
    faculty_id = get_jwt_identity()
    faculty = db.session.get(Faculty, faculty_id)
    if not faculty or not faculty.is_admin:
        return jsonify({'message': 'Unauthorized'}), 403
    data = request.get_json()
    venue = Venue(name=data['name'], location=data['location'], capacity=data['capacity'])
    db.session.add(venue)
    db.session.commit()
    return jsonify({'success': True, 'venue_id': venue.venue_id})

@app.route('/api/venues/<int:venue_id>', methods=['DELETE'])
@jwt_required()
def delete_venue(venue_id):
    faculty_id = get_jwt_identity()
    faculty = db.session.get(Faculty, faculty_id)
    if not faculty or not faculty.is_admin:
        return jsonify({'message': 'Unauthorized'}), 403
    venue = Venue.query.get_or_404(venue_id)
    db.session.delete(venue)
    db.session.commit()
    return jsonify({'success': True})

# Faculty Management (Merged Single Definition)
@app.route('/api/faculty', methods=['GET'])
@jwt_required()
def get_faculty():
    faculty_id = get_jwt_identity()
    faculty = db.session.get(Faculty, faculty_id)
    if not faculty or not faculty.is_admin:
        return jsonify({'message': 'Unauthorized'}), 403

    faculty_list = Faculty.query.all()
    return jsonify([{
        'faculty_id': f.faculty_id,
        'name': f.name,
        'mobile_number': f.mobile_number,
        'email_id': f.email_id,
        'rfid_tag': f.rfid_tag,
        'is_admin': f.is_admin
    } for f in faculty_list])

@app.route('/api/faculty', methods=['POST'])
@jwt_required()
def add_faculty():
    faculty_id = get_jwt_identity()
    faculty = db.session.get(Faculty, faculty_id)
    if not faculty or not faculty.is_admin:
        return jsonify({'message': 'Unauthorized'}), 403
    data = request.get_json()
    faculty = Faculty(
        name=data['name'],
        mobile_number=data['mobile_number'],
        email_id=data['email_id'],
        rfid_tag=data.get('rfid_tag'),  # Include RFID tag
        is_admin=data.get('is_admin', False)
    )
    db.session.add(faculty)
    db.session.commit()
    return jsonify({'success': True, 'faculty_id': faculty.faculty_id})

@app.route('/api/faculty/<int:faculty_id>', methods=['DELETE'])
@jwt_required()
def delete_faculty(faculty_id):
    faculty_id_auth = get_jwt_identity()
    faculty = db.session.get(Faculty, faculty_id_auth)
    if not faculty or not faculty.is_admin:
        return jsonify({'message': 'Unauthorized'}), 403
    faculty_to_delete = Faculty.query.get_or_404(faculty_id)
    db.session.delete(faculty_to_delete)
    db.session.commit()
    return jsonify({'success': True})

# Venue Allocation
@app.route('/api/allocations', methods=['GET'])
@jwt_required()
def get_allocations():
    faculty_id = get_jwt_identity()
    faculty = db.session.get(Faculty, faculty_id)
    db.session.expire_all()

    if faculty.is_admin:
        allocations = VenueAllocation.query.join(Faculty).join(Venue).all()
    else:
        allocations = VenueAllocation.query.join(Faculty).join(Venue).filter(VenueAllocation.faculty_id == faculty_id).all()

    return jsonify([{
        'allocation_id': a.allocation_id,
        'faculty_id': a.faculty_id,
        'faculty_name': a.faculty.name,
        'venue_id': a.venue_id,
        'venue_name': a.venue.name,
        'venue_location': a.venue.location,
        'date': a.date.isoformat(),
        'time_slot': a.time_slot,
        'is_present': bool(Attendance.query.filter_by(allocation_id=a.allocation_id, is_present=True).first())
    } for a in allocations])

@app.route('/api/allocations/generate', methods=['POST'])
@jwt_required()
def generate_allocations():
    faculty_id = get_jwt_identity()
    faculty = db.session.get(Faculty, faculty_id)
    if not faculty or not faculty.is_admin:
        return jsonify({'message': 'Unauthorized'}), 403
    data = request.get_json()
    date = datetime.strptime(data['date'], '%Y-%m-%d').date()
    time_slot = data['time_slot']
    faculty_per_venue = int(data['faculty_per_venue'])

    venues = Venue.query.all()
    faculty_list = Faculty.query.filter_by(is_admin=False).all()

    required = len(venues) * faculty_per_venue
    available = len(faculty_list)
    if available < required:
        return jsonify({
            'message': f'Not enough faculty available: {required} needed, but only {available} available'
        }), 400

    VenueAllocation.query.filter_by(date=date, time_slot=time_slot).delete()
    random.shuffle(faculty_list)
    allocations = []
    faculty_index = 0
    for venue in venues:
        for _ in range(min(faculty_per_venue, len(faculty_list) - faculty_index)):
            allocation = VenueAllocation(
                faculty_id=faculty_list[faculty_index].faculty_id,
                venue_id=venue.venue_id,
                date=date,
                time_slot=time_slot
            )
            allocations.append(allocation)
            faculty_index += 1
        if faculty_index >= len(faculty_list):
            break

    db.session.bulk_save_objects(allocations)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Allocations generated successfully'})

@app.route('/api/bulk-import/faculty', methods=['POST'])
@jwt_required()
def bulk_import_faculty():
    faculty_id = get_jwt_identity()
    faculty = db.session.get(Faculty, faculty_id)
    if not faculty or not faculty.is_admin:
        return jsonify({'message': 'Unauthorized'}), 403

    if 'file' not in request.files:
        return jsonify({'message': 'No file part in the request'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'message': 'Please upload an XLSX file'}), 400

    try:
        workbook = openpyxl.load_workbook(BytesIO(file.read()))
        sheet = workbook.active

        expected_headers = ['faculty_id', 'name', 'mobile_number', 'email_id', 'is_admin', 'rfid_tag']
        headers = [cell.value.lower() if cell.value else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        if not all(h in headers for h in expected_headers):
            return jsonify({'message': 'Invalid Excel format. Required columns: faculty_id, name, mobile_number, email_id, is_admin, rfid_tag'}), 400

        imported_count = 0
        updated_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 6:
                continue

            faculty_id, name, mobile_number, email_id, is_admin, rfid_tag = row
            existing_faculty = db.session.get(Faculty, faculty_id) if faculty_id else None

            if existing_faculty:
                existing_faculty.name = name
                existing_faculty.mobile_number = str(mobile_number)
                existing_faculty.email_id = email_id
                existing_faculty.is_admin = bool(is_admin)
                existing_faculty.rfid_tag = rfid_tag
                updated_count += 1
            else:
                new_faculty = Faculty(
                    faculty_id=faculty_id,
                    name=name,
                    mobile_number=str(mobile_number),
                    email_id=email_id,
                    is_admin=bool(is_admin),
                    rfid_tag=rfid_tag
                )
                db.session.add(new_faculty)
                imported_count += 1

        db.session.commit()
        db.session.expire_all()
        return jsonify({
            'success': True,
            'message': f'Successfully imported {imported_count} new faculty and updated {updated_count} existing faculty'
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error during faculty bulk import: {str(e)}")
        return jsonify({'message': f'Error processing file: {str(e)}'}), 500

@app.route('/api/bulk-import/venues', methods=['POST'])
@jwt_required()
def bulk_import_venues():
    faculty_id = get_jwt_identity()
    faculty = db.session.get(Faculty, faculty_id)
    if not faculty or not faculty.is_admin:
        return jsonify({'message': 'Unauthorized'}), 403

    if 'file' not in request.files:
        return jsonify({'message': 'No file part in the request'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'message': 'Please upload an XLSX file'}), 400

    try:
        workbook = openpyxl.load_workbook(BytesIO(file.read()))
        sheet = workbook.active

        expected_headers = ['venue_id', 'name', 'location', 'capacity']
        headers = [cell.value.lower() if cell.value else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        if not all(h in headers for h in expected_headers):
            return jsonify({'message': 'Invalid Excel format. Required columns: venue_id, name, location, capacity'}), 400

        imported_count = 0
        updated_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 4:
                continue

            venue_id = row[0]
            existing_venue = Venue.query.get(venue_id) if venue_id else None

            if existing_venue:
                existing_venue.name = row[1]
                existing_venue.location = row[2]
                existing_venue.capacity = int(row[3])
                updated_count += 1
            else:
                new_venue = Venue(
                    venue_id=venue_id,
                    name=row[1],
                    location=row[2],
                    capacity=int(row[3])
                )
                db.session.add(new_venue)
                imported_count += 1

        db.session.commit()
        db.session.expire_all()
        return jsonify({
            'success': True,
            'message': f'Successfully imported {imported_count} new venues and updated {updated_count} existing venues'
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error during venue bulk import: {str(e)}")
        return jsonify({'message': f'Error processing file: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)